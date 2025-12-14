import openpyxl

wk = openpyxl.load_workbook("D:\\UDEMY\\TestData.xlsx")
obj = wk['Sheet1']

# print(obj['A3'].value)

#Another approach is making cell object
#Do after line 4
c1=obj.cell(1,2)    #Or we can also pass keyword arguments like c1=obj.cell(column=1,row=3),the order doesn't matter
print(c1.value)
