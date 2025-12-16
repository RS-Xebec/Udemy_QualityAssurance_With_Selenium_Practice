#Import Package
import openpyxl

wk = openpyxl.load_workbook("D:\\UDEMY\\TestData.xlsx")
sh = wk['Sheet1']

#Find rows having data

rows = sh.max_row
column = sh.max_column

print("Total rows are " + str(rows))
print("Total columns are " + str(column))

# for i in range(1,rows+1):           
#     for j in range(1,column+1):
#         c= sh.cell(i,j)
#         print(c.value)

# Another approach is :  do after line 13
for r in sh['A1':'C4']:
    for c in r:
        print(c.value)
